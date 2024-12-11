import os 
from openpyxl import load_workbook

class AutoResult: #pythonからNIPの変数に値（配列）を代入する
    def initialize(self, directory): #おまじない
        os.chdir(path = directory)

    def terminate(self): #おまじない
        pass
    
    def execute(self, solution):
        try:
            tray_info_id = solution.get_variable_id("tray_info") # 配列の中身は{1:トレー番号_現在値、2:ポケット番号_現在値、3：ワーク番号_現在値}
            results_id = solution.get_variable_id("results") # 配列の中身は{1:ズレ、2:濃淡、3：逆印刷、、、、、}
            tray_info = list(solution.get_variable(tray_info_id).values())
            results = list(solution.get_variable(results_id).values())
            print(tray_info, results)

            wb=load_workbook(r'C:\Users\4083625\Downloads\検収ワーク_平松.xlsx')
            sheet=wb.get_sheet_by_name('量試リスト1_大3')

            now = 49

            start_row = (now-41) *52 +4
            start_colum = 16

            for i, result in enumerate(results):

                sheet.cell(row = start_row + tray_info[1] + (tray_info[0]-1)*50, column = start_colum + i + i//(8-1)*2, value = result)
                # トレー番号(tray_info[0])が2のときはrawを50ずらす

            wb.save(r"C:\Users\4083625\Downloads\検収ワーク_平松.xlsx")

            return solution.judge_pass()

        except Exception as  e:
            print(type(e), e)
            return solution.judge_fail()
